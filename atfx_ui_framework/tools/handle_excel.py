'''
Author: Tyler96-QA 1718459369@qq.com
Date: 2023-03-28 14:35:42
LastEditors: TylerQA 990687322@qq.com
LastEditTime: 2023-04-25 21:28:39
FilePath: \Full_stack\handle_excel\handle_excel.py
Description: excel文件读取与写入
'''
import openpyxl
from loguru import logger

class Case(object):
    """
    存储测试数据的类
    """
    def __init__(self,attrs:zip) -> None:
        """
        初始化用例
        :param attrs : zip对象，解包后：[(key,value),(key1,value1)....]
        """
        for item in attrs:
            setattr(self,item[0],item[1]) #set实例对象的属性
        super().__init__()


class HandleExcel(object):

    workbook = None
    """
    对excel操作的封装，读取/写入数据，以及删除sheet表等，
    默认是对Sheet表进行操作
    """

    def __init__(self,filename:str,sheetname:str='Sheet1') -> None:
        """
        初始化读取对象
        :param filename:文件名字
        :param sheetname:sheet表单名
        """
        if not HandleExcel.workbook:
            HandleExcel.workbook = openpyxl.load_workbook(filename)  
        #打开工作簿
        #选择sheet表单
        self.sheet=HandleExcel.workbook[sheetname]
        #定义实例属性
        self.filename=filename
        self.sheetname =sheetname
        # super().__init__()
        #获取excel的最大行
        self.max_rows=self.sheet.max_row

    #读取excel全部内容
    def read_excel_data(self)->list:
        """
        每一条读取出来的数据存储在列表中
        """
        logger.info(f'读取excle文件 {self.filename} 的 {self.sheetname} sheet表单')
        excel_list=[]
        head_list=[]
        #获取表头
        for row in list(self.sheet.rows)[0]:
            head_list.append(row.value)

        if None in head_list:
            raise KeyError('表头为空，请检查excel文档')

        #除去首行，遍历剩下的行数
        for row in list(self.sheet.rows)[1:]:
            data_list=[]
            for cell in row:#每一行中单元格的内容,以元祖/字典的形式添加

                data_list.append(cell.value) #所有数据以字符串的形式添加
            else:
                excel_list.append(dict(zip(head_list,data_list)))
        else:
            HandleExcel.workbook.close()
            return excel_list


    #读取excel全部内容并以对象的形式保存在Case类中
    def read_excel_data_obj(self)->object:
        """
        每一条读取出来的数据作为一个对象存储在列表中
        """
        logger.info(f'读取excle文件 {self.filename} 的 {self.sheetname} sheet表单')
        head_list=[]
        excel_list=[]
        #遍历首行，将首行的字段存在列表中
        for row in list(self.sheet.rows)[0]:
            head_list.append(row.value)

        if None in head_list:
            raise KeyError('表头为空，请检查excel文档')        
        
        #除去首行，遍历剩下的行数
        for row in list(self.sheet.rows)[1:]:
            data_list=[]
            for cell in row:#每一行中单元格的内容,以元祖/字典的形式添加
                data_list.append(cell.value) 
            else:
                #将该条数据存入case对象中
                case_obj=Case(list(zip(head_list,data_list))) #每一次循环下，实例对象的属性都被重新赋值，将打包好的每一行用例作为参数传给init方法，赋值属性
                excel_list.append(case_obj)
        else:
            logger.info(f'返回读取内容：{excel_list}')
            HandleExcel.workbook.close()
            return excel_list
    

    
    #指定行来读取excel内容并储存在对象中
    def read_exceldata_rows_obj(self,row_list:list)->object:

        """
        指定行读取excel内容
        :param row_list:列表:指定行来读取内容 如[1,2]读取第1,2行数据
        """
        logger.info(f'读取excle文件 {self.filename} 的 {self.sheetname} sheet表单')
        head_list=[]
        excel_list=[]
        #遍历首行，将首行的字段存在列表中
        for row in list(self.sheet.rows)[0]:
            head_list.append(row.value)
        
        if None in head_list:
            HandleExcel.workbook.close()
            raise KeyError('表头为空，请检查excel文档')
        

        if len(row_list) == 0:
            return self.read_excel_data_obj()
        else:
            if all(1<=int(i)<=self.max_rows for i in row_list): #判断参数是否正确
                for row in row_list:
                    data_list=[]
                    for cell in list(self.sheet.rows)[row-1]:
                        data_list.append(cell.value)
                    else:
                        #将该条数据存入case对象中
                        case_obj=Case(list(zip(head_list,data_list))) ##将打包好的每一行用例作为参数传给init方法，赋值给实例对象的属性
                        excel_list.append(case_obj)
                else:
                    HandleExcel.workbook.close()
                    return excel_list
            else:
                print('请输入正确的行数')
    
    
    #指定列来读取excel内容并储存在对象中
    def read_exceldata_column_obj(self,col_list:list)->object:
        """
        指定列读取excel内容
        :param col_list:字母，指定列来读取内容 如['A','B']读取第AB列数据
        """
        logger.info(f'读取excle文件 {self.filename} 的 {self.sheetname} sheet表单')
        head_list=[]
        excel_list=[]
        for arg in col_list:
            head_list.append(self.sheet['{}1'.format(arg)].value)
        
        if None in head_list:
            HandleExcel.workbook.close()
            raise KeyError('表头为空，请检查excel文档')
        
        if len(col_list) == 0:
            return self.read_excel_data_obj()
        
        for row in range(2,self.max_rows+1):
            data_list=[]
            for arg in col_list:
                data_list.append(self.sheet['{}{}'.format(arg,row)].value)
            else:
                case_obj=Case(list(zip(head_list,data_list)))
                excel_list.append(case_obj)
        else:
            HandleExcel.workbook.close()
            return excel_list


    #写入数据
    def sava_excel_data(self,cloumn:None,row:int,value:any)->None:
        """
        存储数据到excel
        :param cloumn:列，数字或者英文字母
        :param row:行
        :param value:存储数据
        """
        logger.info(f'写入{self.filename}文件 {self.sheetname} 表单 {cloumn} 列{row} 行：{value}')
        #判断是根据列名还是列数写入数据
        try:
            if isinstance(cloumn,int):
                self.sheet.cell(row,cloumn,value)
            else:
                self.sheet['{}{}'.format(cloumn,row)].value=value
        except BaseException as msg:
            print(f'请勿打开需要写入数据的excel文件：{msg}')
        else:
            HandleExcel.workbook.save(self.filename)
            HandleExcel.workbook.close()
            


if __name__=='__main__':
    from global_SystemEnv import SystemEnv
    import os

    register_data = HandleExcel(os.path.join(SystemEnv.DATA_DIR,'testcases.xlsx'),'gm_register')
    kyc_data = HandleExcel(os.path.join(SystemEnv.DATA_DIR,'testcases.xlsx'),'gm_kyc')
    register_data.sava_excel_data('J',2,'634dd56485456')
    kyc_data.sava_excel_data('J',2,'9877897897ooo')

    # print(len(data))
    # for i in data:
    #     res=requests.request(i.method,i.url,json=eval(i.req_data),files=i.files,data=None)
    #     print(res.text)
    # data=handleData.read_exceldata_rows_obj(2,3)
    # for i in data:
    #     print(i.excepted)
    # import handle_gen_data
    # import handle_log 
    # my = getattr(handle_log,'MyLog')
    # print(my)
